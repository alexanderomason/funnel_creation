#!/usr/bin/env python
# coding: utf-8

# In[1]:


get_ipython().run_line_magic('config', 'Completer.use_jedi = False')
from dotenv import load_dotenv
from dotenv import find_dotenv
import boto3
import sys
import numpy as np

import pandas as pd
from pyathena import connect
from pyathena.cursor import DictCursor

import requests
from pyathena.pandas.cursor import PandasCursor
import datetime
import tempfile
import os
from plotly import graph_objects as go
import kaleido
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import numbers
from openpyxl.styles import Font
load_dotenv(find_dotenv())


# In[41]:


def date_conditions(start_date,end_date):
    '''

    Parameters
    ----------
    start_date : YYYY-MM-DD
    
        DESCRIPTION.
        in str format
        
    end_date : YYYY-MM-DD
        DESCRIPTION.
        in str format
        

    Returns
    -------
    Returns the Presto syntax required to query over the range of dates given the partitions, year, month, day.

    '''
    dates=pd.date_range(start=start_date, end = end_date, freq="D" )
    dates=pd.Series(dates)
    dates=dates.apply(lambda x: x.isoformat().split('T')[0].split('-'))
    dates=pd.DataFrame(list(dates),columns=['year','month','day'])
    start_dates=dates.groupby(['year','month']).first().reset_index()
    end_days=dates.groupby(['year','month']).last().reset_index()['day']
    start_dates['end_day']=end_days
    date_str=''
    for x in start_dates.values:
        date_str+= ' OR '+f'(year = \'{x[0]}\' AND month = \'{x[1]}\' AND (day BETWEEN \'{x[2]}\' AND \'{x[3]}\'))'
    return('('+date_str[4:]+')')





def df_to_png(frame,out_path=None, scale=None):
    '''
    

    Parameters
    ----------
    frame : pd.DataFrame
        DESCRIPTION.
        output of query as dataframe.
        
    out_path : TYPE, optional
        DESCRIPTION. The default is None.

    Returns
    -------
    None, but saves the static funnel image as a png at out_path.

    '''
    
    fig = go.Figure(go.Funnel(
        y = ['Starting Sessions', 'PDP Visits', 'Carts Created', 'Checkout', 'Purchase Complete'],
        x = list((frame.values)[0]),
        textinfo = "value+percent initial",
        textfont = {'size': 18}))
    fig.update_traces(texttemplate="%{value:,d}<br>%{percentInitial:,.1%}")
    
    if scale!= None:
        fig.update_layout(autosize=False, width=int(scale*1200), height=int(scale*600), font=dict(size=18))
    if scale==None:
        fig.update_layout(autosize=False, width=1200, height=600, font=dict(size=18))
        

    
    
    if out_path!=None:
        fig.write_image(out_path)
    else:
        fig.show()

def get_funnel_data(survey_id,start_date,end_date):
    cursor = connect(aws_access_key_id=os.getenv('AWSACCESSKEY'),
                aws_secret_access_key=os.getenv('AWSSECRETKEY'),
                s3_staging_dir=os.getenv('S3_STAGING_DIR'),
                region_name=os.getenv('REGION_NAME')).cursor(PandasCursor)
    infile=open('better_funnel_query.sql')
    
    prestoSql = infile.read()
    infile.close()
    prestoSql=prestoSql.replace('date_con',date_conditions(start_date,end_date))
    prestoSql=prestoSql.replace('replace_id',survey_id)
    
        
    df = cursor.execute(prestoSql,  cache_expiration_time=3600).as_pandas()
    print('got data')
    return(df)
def get_small_frame(df):
        '''creates the small frame to be used in the excel file created below.'''
        # SMALL FRAME
        
        smold={'key_action':['session_starting_timestamp', 'pdp_timestamp',  'in_cart_timestamp', 'checkout_timestamp', 'purchased_timestamp'],'count':list((df.values)[0])}
        tab=pd.DataFrame(smold)
        
        tab['percent'] = tab['count']/tab['count'][0]
        

        tab.key_action=['Starting Sessions', 'PDP Visits', 'Carts Created', 'Checkout', 'Purchase Complete']
    
        tab['% of Previous']=[1,]+list(np.array(tab['count'])[1:]/np.array(tab['count'])[:-1])
        tab.columns=['key_action', 'Count', '% of Initial', '% of Previous']
        small_frame=tab
        never_visit_PDP=1-small_frame['Count'][1]/small_frame['Count'][0]
        never_create_cart=1-small_frame['Count'][2]/small_frame['Count'][1]
        cart_abandon=1-small_frame['Count'][4]/small_frame['Count'][2]
        checkout_abandon=1-small_frame['Count'][4]/small_frame['Count'][3]
        conversion_rate=small_frame['Count'][4]/small_frame['Count'][0]
        key_stats=[never_visit_PDP,never_create_cart,cart_abandon,checkout_abandon,conversion_rate]
        key_stats_description=['Never visit a PDP','Never create a cart','Cart abandonment rate','Checkout abandonment rate','Conversion rate']
        small_frame['Key Stats %']=key_stats
        small_frame['Key Stats']=key_stats_description
        return(small_frame)
def create_sheet(df,survey_id,start_date,end_date):
    '''a lot of openpyxl formatting, a temporary directory is created for the output of df_to_png and is 
        destoyed at the end of the process'''
    now=datetime.datetime.now().isoformat().split('T')[0]
    temporary_dir = tempfile.TemporaryDirectory(dir = os.getcwd())
    png_path=temporary_dir.name+'/'+survey_id+'.png'
    df_to_png(df,png_path)
    small_frame = get_small_frame(df)
        
        
    WB=openpyxl.Workbook()
    ws=WB.worksheets[0]
    ws.append([None])
    ws.append([None])
    ws.append([None,]*3+list(small_frame.columns))

    for row in small_frame.values:
        ws.append([None,]*3+list(row))
            
    ws.title=survey_id
        
    for cell in ws['E4':'E8']:
        cell[0].number_format = '#,##0'
    for cell in ws['F4':'F8']:
        cell[0].number_format = '0.00%'
    for cell in ws['G4':'G8']:
        cell[0].number_format = '0.00%'
    bold=Font(bold=True) 
    bold_cells=['D3','E3','F3','G3','H3','I3']
    for x in bold_cells:
        ws[x].font=bold
        ws.column_dimensions[x[0]].width = 20
            

    ws['H4'].number_format = '0.00%'

    ws['H5'].number_format = '0.00%'

    ws['H6'].number_format = '0.00%'

    ws['H7'].number_format = '0.00%'

    ws['H8'].number_format = '0.00%'

        
    ws['D2'].value='Start Date'
    ws['D2'].font=bold
        
    ws['E2'].value=start_date
        
    ws['F2'].value='End Date'
    ws['F2'].font=bold
        
    ws['G2'].value=end_date
        
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
        
        
    img = openpyxl.drawing.image.Image(png_path)
        
    img.anchor = 'D12'
        
    ws.add_image(img)
        
    out_path=survey_id.replace('.','_')
    out_path+='-'+start_date.replace('-','')
    out_path+='-'+end_date.replace('-','')
    out_path+=' - '+now.replace('-','')+'.xlsx'
        
    
    WB.save(out_path)
            
        
    temporary_dir.cleanup()

def funnel_creator(survey_id,start_date,end_date):
    df = get_funnel_data(survey_id,start_date,end_date)
    small_frame=get_small_frame(df)
    create_sheet(df,survey_id,start_date,end_date)


# In[44]:


#funnel_data=get_funnel_data('prod.com.jcpenney.acconfig', '2023-01-01', '2023-01-03')


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[46]:


#funnel_creator('prod.com.jcpenney.acconfig', '2023-01-01', '2023-01-03')


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:




